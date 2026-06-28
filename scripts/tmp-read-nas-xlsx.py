import json
import openpyxl

path = r"C:\tmp\NAS管理台帳\NAS一覧.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["NAS一覧"]

headers = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(8, c).value
    if v:
        headers[c] = str(v).replace("\n", "")

print("HEADERS R8:")
for c, h in headers.items():
    print(f"  {openpyxl.utils.get_column_letter(c)}: {h}")

col_map = {c: headers[c] for c in headers}
records = []
for r in range(9, ws.max_row + 1):
    row = {}
    has_data = False
    for c, h in col_map.items():
        v = ws.cell(r, c).value
        if v is not None and str(v).strip():
            has_data = True
            row[h] = str(v).strip()
    if has_data and any(k in row for k in ["IPアドレス", "ステータス", "拠点名"]):
        row["_row"] = r
        records.append(row)

print(f"\nRECORD COUNT: {len(records)}")
for rec in records:
    print(f"--- R{rec['_row']} ---")
    for k, v in rec.items():
        if k == "_row":
            continue
        vv = v.replace("\r\n", " / ").replace("\n", " / ")
        if len(vv) > 60:
            vv = vv[:60] + "..."
        print(f"  {k}: {vv}")

out = {
    "version": "2026-06-28-live-excel",
    "headerRow": 8,
    "recordCount": len(records),
    "records": records,
}
out_path = "docs/plans/tmp-nas-xlsx-structure.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)
print(f"\nWrote {out_path}")
