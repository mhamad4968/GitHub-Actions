#!/usr/bin/env python3
"""JMA obsdl CSV + Excel → workdays-5yr-omiya.json（全閾値ブロック）"""
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
JSON_PATH = ROOT / "scripts/data/workdays-5yr-omiya.json"
EXCEL_PATH = Path(r"C:\tmp\稼働日数算出ツール\稼働日数算出ツール20260613.xlsx")
DEFAULT_FOLDER = Path(r"C:\tmp\稼働日数算出ツール")

WIND_THRESHOLDS = [10, 15, 20, 30]
RAIN_THRESHOLDS = [1, 10, 30, 50, 70, 100]
WIND_COLS = {10: 4, 15: 5, 20: 6, 30: 7}
RAIN_COLS = {1: 4, 10: 5, 30: 6, 50: 7, 70: 8, 100: 9}


def read_csv_text(path: Path) -> str:
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "cp932", "shift_jis"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def parse_jma_daily_csv(path: Path):
    text = read_csv_text(path)
    date_re = re.compile(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})")
    num_re = re.compile(r"[^\d.-]")
    rows = []
    for line in text.splitlines():
        line = line.strip()
        if not date_re.match(line):
            continue
        parts = line.split(",")
        if len(parts) < 2:
            continue
        y, mo, d = date_re.match(line).groups()
        try:
            val = float(num_re.sub("", parts[1].strip()) or "nan")
        except ValueError:
            continue
        if val != val:
            continue
        rows.append(
            {
                "date": f"{int(y):04d}-{int(mo):02d}-{int(d):02d}",
                "value": val,
            }
        )
    return rows


def counts_from_daily(rows, threshold):
    counts = defaultdict(lambda: defaultdict(int))
    for row in rows:
        y, mo, _ = row["date"].split("-")
        counts[y][int(mo)]  # 全期間を0で初期化（該当日0件の閾値でも年を登録）
    for row in rows:
        if row["value"] < threshold:
            continue
        y, mo, _ = row["date"].split("-")
        counts[y][int(mo)] += 1
    return counts


def block_from_counts(counts, label):
    years = sorted(counts.keys())
    months = []
    for m in range(1, 13):
        by_year = {y: int(counts[y].get(m, 0)) for y in years}
        avg = sum(by_year.values()) / len(by_year) if years else 0
        months.append({"m": m, "byYear": by_year, "avg": avg})
    return {"label": label, "years": years, "months": months}


def extract_excel_blocks(wb, sheet_name, col_map, prefix, unit):
    ws = wb[sheet_name]
    blocks = {}
    for th, col in col_map.items():
        key = f"{prefix}_ge{th}_{unit}"
        by_year = {}
        for r in range(1, 120):
            y = ws.cell(r, 2).value
            if isinstance(y, (int, float)) and 2000 < y < 2100:
                y = str(int(y))
                months = {}
                for rr in range(r, r + 15):
                    m = ws.cell(rr, 3).value
                    if isinstance(m, (int, float)) and 1 <= m <= 12:
                        v = ws.cell(rr, col).value
                        if v is not None:
                            months[int(m)] = int(v)
                if months:
                    by_year[y] = months
        counts = defaultdict(lambda: defaultdict(int))
        for y, months in by_year.items():
            for m, v in months.items():
                counts[y][m] = v
        blocks[key] = block_from_counts(counts, f">={th}{unit.replace('_mm', 'mm').replace('_ms', 'm/s')}")
    return blocks


def merge_block(existing, csv_counts):
    if not existing:
        existing = {"months": [], "years": []}
    years = sorted(set(existing.get("years", [])) | set(csv_counts.keys()))
    by_m = {row["m"]: row for row in existing.get("months", [])}
    for m in range(1, 13):
        row = by_m.get(m, {"m": m, "byYear": {}})
        by_year = dict(row.get("byYear", {}))
        for y in csv_counts:
            by_year[y] = int(csv_counts[y].get(m, 0))
        vals = [by_year.get(y, 0) for y in years]
        by_m[m] = {"m": m, "byYear": by_year, "avg": sum(vals) / len(vals) if vals else 0}
    existing["months"] = [by_m[m] for m in range(1, 13)]
    existing["years"] = years
    return existing


def main():
    folder = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FOLDER
    wind_csv = folder / "風速.csv"
    rain_csv = folder / "降雨.csv"

    data = {
        "location": "大宮地区",
        "windPeriod": "",
        "rainPeriod": "",
        "updated": "2026-06-13",
        "csvSource": str(wind_csv.name) + ", " + str(rain_csv.name),
    }

    if EXCEL_PATH.exists():
        import openpyxl

        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        data.update(extract_excel_blocks(wb, "過去5年月別風速日数", WIND_COLS, "wind", "ms"))
        data.update(extract_excel_blocks(wb, "過去5年月別降雨日数", RAIN_COLS, "rain", "mm"))

    if wind_csv.exists():
        wind_rows = parse_jma_daily_csv(wind_csv)
        for th in WIND_THRESHOLDS:
            key = f"wind_ge{th}_ms"
            counts = counts_from_daily(wind_rows, th)
            data[key] = merge_block(data.get(key), counts)

    if rain_csv.exists():
        rain_rows = parse_jma_daily_csv(rain_csv)
        for th in RAIN_THRESHOLDS:
            key = f"rain_ge{th}_mm"
            counts = counts_from_daily(rain_rows, th)
            data[key] = merge_block(data.get(key), counts)

    wind_years = data.get("wind_ge10_ms", {}).get("years", [])
    rain_years = data.get("rain_ge10_mm", {}).get("years", [])
    if wind_years:
        data["windPeriod"] = f"{wind_years[0]}〜{wind_years[-1]}"
    if rain_years:
        data["rainPeriod"] = f"{rain_years[0]}〜{rain_years[-1]}"

    JSON_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "wind_keys": [k for k in data if k.startswith("wind_")],
                "rain_keys": [k for k in data if k.startswith("rain_")],
                "wind_years": wind_years,
                "rain_years": rain_years,
            },
            ensure_ascii=False,
        )
    )
    print(f"Updated {JSON_PATH}")


if __name__ == "__main__":
    main()
