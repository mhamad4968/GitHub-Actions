#!/usr/bin/env python3
"""
旧フォーマット（Excel）→ kintone 入力アプリ 677 初回投入。
正本: templates/yojitsu-budget-lite/docs/yojitsu-migration-kyu-to-kintone.md §1・§3・§4

前提:
  - シート名「旧フォーマット」、ヘッダ行 2、データ行 3〜、月列 I〜T＝5月〜4月、U＝都度→イニシャル費用（変動費）
  - 月次「予算」のみ移行。実績・予算修正は 0／空。支払内訳は付けない（旧に無い）
  - learning_fixed_budget: 月次予算（I〜T）の合計をミラー（678 running 表示用）。旧 U→initial_variable_budget

使い方:
  python3 scripts/yojitsu-import-2026-budget-xlsx-to-677.py --dry-run
  python3 scripts/yojitsu-import-2026-budget-xlsx-to-677.py --apply --force   # 677 に既存レコードがあっても投入
  python3 scripts/yojitsu-import-2026-budget-xlsx-to-677.py --apply           # 既存0件のときのみ投入
"""
from __future__ import annotations

import argparse
import base64
import json
import sys
from datetime import date, datetime
from pathlib import Path
from urllib.error import HTTPError
from urllib.request import Request, urlopen

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl が必要です。", file=sys.stderr)
    sys.exit(2)

APP_ID = 677
FISCAL_LABELS = ["5", "6", "7", "8", "9", "10", "11", "12", "1", "2", "3", "4"]
MONTH_COL_START = 9  # I = 5月
DEFAULT_XLSX = Path("/mnt/c/tmp/予算管理/2026年度システム推進室_年間予算案20260123.xlsx")


def load_env(repo_root: Path) -> dict[str, str]:
    p = repo_root / ".env"
    if not p.is_file():
        print(f"Missing {p}", file=sys.stderr)
        sys.exit(1)
    env: dict[str, str] = {}
    for line in p.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#") or "=" not in s:
            continue
        k, _, v = s.partition("=")
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        env[k] = v
    return env


def num(x) -> float:
    if x is None or x == "":
        return 0.0
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def code_str(d) -> str:
    if d is None or d == "":
        return ""
    if isinstance(d, float) and d == int(d):
        return str(int(d))
    return str(d).strip()


def cost_category_for_row(month_vals: list, uval) -> str:
    msum = sum(num(x) for x in month_vals)
    un = num(uval)
    if msum == 0 and un > 0:
        return "変動費"
    if msum > 0 and un == 0:
        return "固定費"
    if msum > 0 and un > 0:
        return "固定費"
    return "その他"


def build_monthly_rows(month_vals: list) -> list[dict]:
    rows = []
    for lab, cell in zip(FISCAL_LABELS, month_vals):
        v = num(cell)
        rows.append(
            {
                "value": {
                    "fiscal_month": {"value": lab},
                    "month_budget": {"value": str(int(v)) if v == int(v) else str(v)} if v else {"value": ""},
                    "month_actual": {"value": ""},
                    "month_budget_revision": {"value": "0"},
                }
            }
        )
    return rows


def cell_date_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.isoformat()
    return None


def build_record(
    *,
    row_no: int,
    work_name: str,
    work_code: str,
    summary: str,
    company,
    issue_cell,
    cash_cell,
    month_vals: list,
    uval,
    legacy_b,
) -> dict:
    cc = cost_category_for_row(month_vals, uval)
    msum = sum(num(x) for x in month_vals)
    un = num(uval)
    # 旧 I〜T は月次「予算」のみ。レコード直下の learning_fixed_budget は 678 の running 集計用に
    # 「暦月予算の合計」をミラーする（旧シートにランニング専用列は無いが、ダッシュは lb を参照する）。
    learning_str = ""
    if msum > 0:
        learning_str = str(int(msum)) if msum == int(msum) else str(msum)
    initial_str = ""
    if un > 0:
        initial_str = str(int(un)) if un == int(un) else str(un)
    rec: dict = {
        "work_type_name": {"value": work_name[:255]},
        "work_type_code": {"value": (work_code or "")[:255]},
        "cost_category": {"value": cc},
        "summary_text": {"value": (summary or "").strip()[:10000]},
        "partner_company": {"value": (str(company).strip() if company else "")[:255]},
        "initial_variable_budget": {"value": initial_str} if initial_str else {"value": ""},
        "learning_fixed_budget": {"value": learning_str} if learning_str else {"value": ""},
        "monthly_breakdown": {"value": build_monthly_rows(month_vals)},
        "display_order": {"value": str(int(legacy_b))} if legacy_b is not None else {"value": str(row_no)},
        "legacy_row_no": {"value": str(int(legacy_b))} if legacy_b is not None else {"value": ""},
    }
    idt = cell_date_str(issue_cell)
    if idt:
        rec["issue_date"] = {"value": idt}
    cdt = cell_date_str(cash_cell)
    if cdt:
        rec["legacy_cash_date"] = {"value": cdt}
    extra = []
    if issue_cell is not None and not idt:
        extra.append(f"起票セル: {issue_cell}")
    if cash_cell is not None and not cdt:
        extra.append(f"出納セル: {cash_cell}")
    if extra:
        rec["notes"] = {"value": "旧フォーマット移行メモ（" + " / ".join(extra) + "）"[:10000]}
    return rec


def iter_source_records(ws):
    last_name, last_code = "", ""
    for r in range(3, ws.max_row + 1):
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value
        g = ws.cell(row=r, column=7).value
        h = ws.cell(row=r, column=8).value
        b = ws.cell(row=r, column=2).value
        months = [ws.cell(row=r, column=col).value for col in range(MONTH_COL_START, MONTH_COL_START + 12)]
        uval = ws.cell(row=r, column=21).value
        if c is not None and str(c).strip() == "総計":
            continue
        if c is not None and str(c).strip() and str(c).strip() != "\u3000":
            last_name = str(c).strip()
        if d is not None and d != "":
            last_code = code_str(d)
        msum = sum(num(x) for x in months)
        unum = num(uval)
        es = (e or "").strip() if isinstance(e, str) else (e or "")
        if (not es) and msum == 0 and unum == 0:
            continue
        if not last_name or last_name == "総計":
            continue
        if not es:
            continue
        yield r, last_name, last_code, e, f, g, h, months, uval, b


def kintone_headers(env: dict[str, str]) -> dict[str, str]:
    base = env.get("KINTONE_BASE_URL", "").rstrip("/").removesuffix("/k")
    user = env.get("KINTONE_USERNAME", "")
    pw = env.get("KINTONE_PASSWORD", "")
    if not base or not user or not pw:
        print("KINTONE_BASE_URL, KINTONE_USERNAME, KINTONE_PASSWORD が .env に必要です。", file=sys.stderr)
        sys.exit(1)
    token = base64.b64encode(f"{user}:{pw}".encode()).decode()
    h = {"X-Cybozu-Authorization": token, "Content-Type": "application/json"}
    if env.get("KINTONE_BASIC_AUTH_USERNAME") and env.get("KINTONE_BASIC_AUTH_PASSWORD"):
        bu = env["KINTONE_BASIC_AUTH_USERNAME"]
        bp = env["KINTONE_BASIC_AUTH_PASSWORD"]
        h["Authorization"] = "Basic " + base64.b64encode(f"{bu}:{bp}".encode()).decode()
    return h, base


def http_json(method: str, url: str, headers: dict, body=None):
    data = None if body is None else json.dumps(body).encode("utf-8")
    req = Request(url, data=data, headers=headers, method=method)
    with urlopen(req, timeout=120) as resp:
        return json.loads(resp.read().decode("utf-8"))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--force", action="store_true")
    args = ap.parse_args()
    repo = Path(__file__).resolve().parents[1]
    env = load_env(repo)
    headers, base = kintone_headers(env)

    if not args.xlsx.is_file():
        print(f"ファイルがありません: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    if "旧フォーマット" not in wb.sheetnames:
        print("シート「旧フォーマット」がありません:", wb.sheetnames, file=sys.stderr)
        sys.exit(1)
    ws = wb["旧フォーマット"]

    records: list[dict] = []
    for r, wn, wc, e, f, g, h, months, uval, b in iter_source_records(ws):
        summary = e if isinstance(e, str) else (str(e) if e is not None else "")
        rec = build_record(
            row_no=r,
            work_name=wn,
            work_code=wc,
            summary=summary,
            company=f,
            issue_cell=g,
            cash_cell=h,
            month_vals=months,
            uval=uval,
            legacy_b=b,
        )
        records.append(rec)

    print(f"生成レコード数: {len(records)}（元シート 旧フォーマット）")

    if args.dry_run or not args.apply:
        for i, rec in enumerate(records[:3]):
            print("--- sample", i + 1, "---")
            print(json.dumps(rec, ensure_ascii=False, indent=2)[:2500])
        if not args.apply:
            print("\n本投入: --apply （677 が空のとき）または --apply --force")
        return

    url_count = f"{base}/k/v1/records.json?app={APP_ID}&query=limit%201&totalCount=true"
    try:
        j = http_json("GET", url_count, {k: v for k, v in headers.items() if k != "Content-Type"})
        total = int(j.get("totalCount", 0))
    except HTTPError as ex:
        print("件数取得失敗", ex.read().decode()[:800], file=sys.stderr)
        sys.exit(1)

    if total > 0 and not args.force:
        print(f"677 に既に {total} 件あります。二重投入を避けるため中止しました。--force で続行。", file=sys.stderr)
        sys.exit(1)

    url_post = f"{base}/k/v1/records.json"
    body = {"app": APP_ID, "records": records}
    try:
        out = http_json("POST", url_post, headers, body)
    except HTTPError as ex:
        err = ex.read().decode()
        print("POST 失敗:", err[:4000], file=sys.stderr)
        sys.exit(1)
    ids = out.get("ids")
    if not isinstance(ids, list):
        ids = [x.get("id") for x in out.get("records", []) if isinstance(x, dict)]
    print("投入成功。件数:", len(ids), "ids 先頭10:", ids[:10] if ids else out)


if __name__ == "__main__":
    main()
