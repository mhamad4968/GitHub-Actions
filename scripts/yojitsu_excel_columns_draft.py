#!/usr/bin/env python3
"""
Excel「新フォーマット」シートの行 2〜4 を Markdown 表にダンプする（SPEC / kintone フィールド案のたたき台）。

依存: pip install openpyxl

用法:
  python3 scripts/yojitsu_excel_columns_draft.py [xlsxパス]
  npm run yojitsu:excel-draft

既定 xlsx: WSL なら /mnt/c/tmp/予算管理/2026年度システム推進室_年間予算案20260123.xlsx
環境変数 YOJITSU_EXCEL で上書き可。
"""
from __future__ import annotations

import os
import sys
from pathlib import Path

DEFAULT_WSL = Path("/mnt/c/tmp/予算管理/2026年度システム推進室_年間予算案20260123.xlsx")
SHEET = "新フォーマット"
MAX_COL = 60
ROWS = (2, 3, 4)


def col_letter(n: int) -> str:
    s = ""
    x = n
    while x:
        x, r = divmod(x - 1, 26)
        s = chr(65 + r) + s
    return s


def esc_cell(v) -> str:
    if v is None:
        return ""
    t = str(v).replace("\r", "").replace("\n", " ").replace("|", "\\|").strip()
    return t[:120]


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("openpyxl が必要です: pip install openpyxl", file=sys.stderr)
        return 2

    raw = sys.argv[1] if len(sys.argv) > 1 else os.environ.get("YOJITSU_EXCEL", "")
    path = Path(raw) if raw else DEFAULT_WSL
    if not path.is_file():
        print(f"ファイルがありません: {path}", file=sys.stderr)
        print("第1引数で xlsx を指定するか YOJITSU_EXCEL=... を設定してください。", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        print(f"シート '{SHEET}' がありません: {wb.sheetnames}", file=sys.stderr)
        return 2
    ws = wb[SHEET]
    ncols = min(MAX_COL, ws.max_column or MAX_COL)

    print("# 列見出しドラフト（自動生成）\n")
    print(f"- ソース: `{path}` シート `{SHEET}`\n")

    for r in ROWS:
        letters = [col_letter(c) for c in range(1, ncols + 1)]
        vals = [esc_cell(ws.cell(r, c).value) for c in range(1, ncols + 1)]
        print(f"## 行 {r}\n")
        print("| " + " | ".join(letters) + " |")
        print("| " + " | ".join("---" for _ in letters) + " |")
        print("| " + " | ".join(vals) + " |")
        print()

    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
